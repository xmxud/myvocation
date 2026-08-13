"""Daily executions routes."""
import io
import json
import os
import re
from datetime import date as date_cls, datetime, time as time_cls, timedelta

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import FileResponse, StreamingResponse
from app.database import get_db, query, query_one, execute
from app.models import success, error
from app.core.auth import get_optional_user
from app.services.llm import chat_json, LLMError
from app.services.oss import oss_storage, OSSError

router = APIRouter(prefix="/api/daily-executions", tags=["每日执行"])


@router.get("/{node_id}")
async def list_executions(node_id: int, phase_id: int = None, date: str = None,
                          start_date: str = None, end_date: str = None,
                          focus_id: int = None):
    db = await get_db()
    try:
        # node_id 通常为 THEME/FOCUS_ITEM；按节点子树查询，
        # 使挂在子节点（如重点项）上的执行记录也能被列出
        sql = """WITH RECURSIVE subtree AS (
                   SELECT id FROM planning_nodes WHERE id = ?
                   UNION ALL
                   SELECT n.id FROM planning_nodes n
                   JOIN subtree s ON n.parent_id = s.id
                 )
                 SELECT * FROM daily_executions
                 WHERE node_id IN (SELECT id FROM subtree)"""
        params = [node_id]
        if phase_id:
            sql += " AND phase_id = ?"
            params.append(phase_id)
        if date:
            sql += " AND execution_date = ?"
            params.append(date)
        if start_date:
            sql += " AND execution_date >= ?"
            params.append(start_date)
        if end_date:
            sql += " AND execution_date <= ?"
            params.append(end_date)
        # 重点科目过滤：精确匹配记录所属的重点项节点
        if focus_id:
            sql += " AND node_id = ?"
            params.append(focus_id)
        sql += " ORDER BY execution_date DESC, planned_start_time, id DESC"
        rows = await query(db, sql, tuple(params))
        return success(rows)
    finally:
        await db.close()


@router.post("/upload-attachment")
async def upload_attachment(file: UploadFile = File(...), user=Depends(get_optional_user)):
    """上传执行记录附件到阿里云 OSS，返回 {key, url, name}。"""
    data = await file.read()
    if not data:
        raise HTTPException(400, "附件内容为空")
    if len(data) > 20 * 1024 * 1024:
        raise HTTPException(400, "附件大小不能超过 20MB")
    try:
        info = oss_storage.upload_attachment(data, file.filename or "attachment",
                                             prefix="execution-docs")
    except OSSError as e:
        raise HTTPException(502, str(e))
    return success({**info, "name": file.filename}, "附件上传成功")


@router.post("/ai-generate")
async def ai_generate_executions(body: dict, user=Depends(get_optional_user)):
    """根据阶段行动指南（含父阶段）调用千问生成该阶段每一天的执行任务。

    幂等策略：先清除该阶段下 source='auto' 的旧记录，再写入新生成的任务。
    """
    phase_id = body.get("phase_id")
    if not phase_id:
        raise HTTPException(400, "缺少 phase_id")

    db = await get_db()
    try:
        phase = await query_one(db, "SELECT * FROM phases_v2 WHERE id = ?", (phase_id,))
        if not phase:
            raise HTTPException(404, "阶段不存在")

        # 收集本阶段及全部父阶段的行动指南（point_type='action'）
        chain = []  # [(phase_row, [action_points])]，0=本阶段，之后依次为父/祖父阶段
        cur = phase
        while cur:
            pts = await query(
                db,
                "SELECT * FROM phase_points WHERE phase_id = ? AND point_type = 'action'"
                " ORDER BY sort_order, id",
                (cur["id"],),
            )
            chain.append((cur, pts))
            if cur["parent_id"]:
                cur = await query_one(db, "SELECT * FROM phases_v2 WHERE id = ?", (cur["parent_id"],))
            else:
                cur = None

        if not any(pts for _, pts in chain):
            raise HTTPException(400, "该阶段及其父阶段均无行动指南条目，无法生成")

        # 加载该主题下的科目（FOCUS_ITEM 子节点），用于把任务关联到对应科目
        subjects = await query(
            db,
            "SELECT id, title FROM planning_nodes"
            " WHERE parent_id = ? AND node_type = 'FOCUS_ITEM' ORDER BY id",
            (phase["node_id"],),
        )
        subject_lines = "\n".join(f"- id={s['id']}：{s['title']}" for s in subjects) or "（无）"

        # 日期范围
        try:
            start = date_cls.fromisoformat(phase["start_date"])
            end = date_cls.fromisoformat(phase["end_date"])
        except (TypeError, ValueError):
            raise HTTPException(400, "阶段日期格式无效")
        if end < start:
            raise HTTPException(400, "阶段日期范围无效")
        days = [start + timedelta(days=i) for i in range((end - start).days + 1)]
        if len(days) > 62:
            raise HTTPException(400, "阶段跨度超过 62 天，请先拆分阶段再生成")

        # 组装提示词
        weekday_cn = "一二三四五六日"
        date_list = "、".join(
            f"{d.isoformat()}(周{weekday_cn[d.weekday()]})" for d in days
        )

        def fmt_actions(pts):
            lines = []
            for p in pts:
                extra = {}
                if p["extra_data"]:
                    try:
                        extra = json.loads(p["extra_data"])
                    except (json.JSONDecodeError, TypeError):
                        extra = {}
                freq = {"daily": "每日", "weekly": "每周", "once": "一次"}.get(
                    extra.get("frequency"), extra.get("frequency") or ""
                )
                minutes = extra.get("estimated_minutes")
                suffix = ""
                if freq or minutes:
                    suffix = f"（{freq}{('，约' + str(minutes) + '分钟') if minutes else ''}）"
                lines.append(f"- {p['content']}{suffix}")
            return "\n".join(lines) if lines else "（无）"

        own_actions = fmt_actions(chain[0][1])
        parent_actions = "\n".join(
            f"【父阶段 {p['phase_number']}. {p['title']}】\n{fmt_actions(pts)}"
            for p, pts in chain[1:]
        ) or "（无）"

        messages = [
            {"role": "system", "content":
             "你是一名学习计划助手，擅长把阶段行动指南拆解为每天可执行的具体任务。"
             "只输出 JSON，不要输出任何解释性文字。"},
            {"role": "user", "content": f"""阶段：{phase['phase_number']}. {phase['title']}
日期范围：{phase['start_date']} 至 {phase['end_date']}

本阶段行动指南：
{own_actions}

父阶段行动指南（在整个阶段期间持续生效，也需要安排进每天）：
{parent_actions}

科目列表（任务必须关联到对应科目，focus_id 取这里的 id）：
{subject_lines}

请为该阶段的每一天生成 6-10 项执行任务，要求：
1. 每条任务必须根据内容关联到对应科目，输出 focus_id（取自上方科目列表的 id）；实在无法归类的用 null
2. 覆盖上述全部行动指南，且要顾及所有科目：每天安排不同科目的任务，整个阶段内每个科目都应被安排到
3. 频次为"每日"的行动指南，每天必须安排在相同的固定时间（planned_start_time 逐日保持一致）
4. 每日任务合理分配在 06:30-23:30 之间，时间不重叠，中午 12:00-13:30 可安排休息/午餐，晚上 22:00-23:30 之间安排复盘总结
5. title 为简洁的任务描述（含具体内容，如"数学每日一练：函数专项 10 题"）
6. planned_start_time 格式 "HH:MM"，planned_duration 为整数分钟
7. date 必须严格取自以下日期：{date_list}
8. 每周可安排 1 天任务量稍轻的总结/复盘日，需要分时间段安排每门科目的复盘

输出 JSON 格式：
{{"days": [{{"date": "YYYY-MM-DD", "tasks": [{{"title": "...", "focus_id": 12, "planned_start_time": "08:00", "planned_duration": 30}}]}}]}}"""},
        ]

        try:
            result = await chat_json(messages)
        except LLMError as e:
            raise HTTPException(502, str(e))

        # 解析并校验生成结果
        valid_dates = {d.isoformat() for d in days}
        valid_focus = {s["id"] for s in subjects}
        items = []
        for day in result.get("days", []):
            dt = day.get("date")
            if dt not in valid_dates:
                continue
            for t in (day.get("tasks") or [])[:8]:
                title = str(t.get("title") or "").strip()
                if not title:
                    continue
                try:
                    duration = int(t["planned_duration"]) if t.get("planned_duration") is not None else None
                except (TypeError, ValueError):
                    duration = None
                start_time = t.get("planned_start_time")
                if not (isinstance(start_time, str) and len(start_time) == 5 and ":" in start_time):
                    start_time = None
                # 科目关联：仅接受科目列表内的 id，否则回退到阶段所属节点
                try:
                    focus_id = int(t["focus_id"]) if t.get("focus_id") is not None else None
                except (TypeError, ValueError):
                    focus_id = None
                if focus_id not in valid_focus:
                    focus_id = None
                items.append((dt, focus_id or phase["node_id"], title, start_time, duration))

        if not items:
            raise HTTPException(502, "AI 未生成任何有效任务，请重试")

        # 清除该阶段旧的自动生成记录，写入新任务
        await execute(db, "DELETE FROM daily_executions WHERE phase_id = ? AND source = 'auto'", (phase_id,))
        uid = phase["user_id"] or (user["user_id"] if user else 1)
        for dt, node_id, title, start_time, duration in items:
            await execute(
                db,
                """INSERT INTO daily_executions
                   (node_id, phase_id, user_id, execution_date, title,
                    planned_start_time, planned_duration, source,
                    is_done, completion_percent, result_score)
                   VALUES (?, ?, ?, ?, ?, ?, ?, 'auto', 0, 0, 0)""",
                (node_id, phase_id, uid, dt, title, start_time, duration),
            )

        return success(
            {"phase_id": phase_id, "days": len({i[0] for i in items}), "created": len(items)},
            f"已生成 {len(items)} 项任务",
        )
    finally:
        await db.close()


@router.post("")
async def create_execution(body: dict, user=Depends(get_optional_user)):
    db = await get_db()
    try:
        uid = user["user_id"] if user else 1
        last_id, _ = await execute(
            db,
            """INSERT INTO daily_executions
               (node_id, phase_id, user_id, execution_date, title,
                planned_start_time, planned_duration, source, source_point_id,
                is_done, completion_percent, notes, duration_minutes, result_score,
                actual_start_time, actual_end_time, attachments)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (body.get("node_id"), body.get("phase_id"), uid,
             body.get("execution_date", ""), body.get("title"),
             body.get("planned_start_time"), body.get("planned_duration"),
             body.get("source", "manual"), body.get("source_point_id"),
             body.get("is_done", 0), body.get("completion_percent", 0),
             body.get("notes"), body.get("duration_minutes"), body.get("result_score", 0),
             body.get("actual_start_time"), body.get("actual_end_time"),
             body.get("attachments")),
        )
        row = await query_one(db, "SELECT * FROM daily_executions WHERE id = ?", (last_id,))
        return success(row, "执行记录创建成功")
    finally:
        await db.close()


@router.put("/{exec_id}")
async def update_execution(exec_id: int, body: dict):
    db = await get_db()
    try:
        fields = []
        values = []
        for key in ["title", "planned_start_time", "planned_duration", "is_done",
                     "completion_percent", "duration_minutes", "notes", "result_score",
                     "mood", "execution_date", "node_id", "phase_id",
                     "actual_start_time", "actual_end_time", "attachments"]:
            if key in body and body[key] is not None:
                fields.append(f"{key} = ?")
                values.append(body[key])
        if not fields:
            return success(await query_one(db, "SELECT * FROM daily_executions WHERE id = ?", (exec_id,)))
        fields.append("updated_at = CURRENT_TIMESTAMP")
        values.append(exec_id)
        await execute(db, f"UPDATE daily_executions SET {', '.join(fields)} WHERE id = ?", tuple(values))
        return success(await query_one(db, "SELECT * FROM daily_executions WHERE id = ?", (exec_id,)), "更新成功")
    finally:
        await db.close()


@router.delete("")
async def clear_phase_executions(phase_id: int = None):
    """清空某阶段下的全部执行任务（用于"清空阶段计划"）。"""
    if not phase_id:
        raise HTTPException(400, "缺少 phase_id")
    db = await get_db()
    try:
        _, changes = await execute(db, "DELETE FROM daily_executions WHERE phase_id = ?", (phase_id,))
        return success({"phase_id": phase_id, "deleted": changes}, f"已清空 {changes} 项任务")
    finally:
        await db.close()


@router.delete("/{exec_id}")
async def delete_execution(exec_id: int):
    db = await get_db()
    try:
        await execute(db, "DELETE FROM daily_executions WHERE id = ?", (exec_id,))
        return success({"deleted_id": exec_id}, "删除成功")
    finally:
        await db.close()


# ── Excel 计划导入 ──

@router.get("/plan-template/download")
async def download_plan_template():
    """下载学习计划导入模版"""
    root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__))))
    path = os.path.join(root, "docs", "design", "学习计划模版.xlsx")
    return FileResponse(path, filename="学习计划模版.xlsx",
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def _sheet_phase_number(title: str):
    """从页签名提取阶段编号，如 '1.2. 暑假第二周（8.8-8.15）' -> '1.2'。"""
    m = re.match(r"^\s*(\d+(?:\.\d+)*)\s*[\.、]?\s*\S", title or "")
    return m.group(1) if m else None


def _parse_header_date(value, phase_start, phase_end):
    """从列表头（如 '7月31日\\n周五' 或日期单元格）解析出具体日期，年份取阶段所在年份。"""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date_cls):
        return value.isoformat()
    m = re.search(r"(\d{1,2})\s*月\s*(\d{1,2})\s*日", str(value or ""))
    if not m:
        return None
    month, day = int(m.group(1)), int(m.group(2))
    try:
        cand = date_cls(phase_start.year, month, day)
    except ValueError:
        return None
    # 跨年阶段（如 12月-1月）按与阶段范围的偏差修正年份
    if (phase_start - cand).days > 180:
        cand = date_cls(phase_start.year + 1, month, day)
    elif (cand - phase_end).days > 180:
        cand = date_cls(phase_start.year - 1, month, day)
    return cand.isoformat()


def _parse_time_label(value):
    """从行首时间段（如 '9:00-11:00'、'23:30'、'20:50-21:10\\n回家'）解析 (HH:MM, 时长分钟)。"""
    if isinstance(value, datetime):
        return value.strftime("%H:%M"), None
    if isinstance(value, time_cls):
        return value.strftime("%H:%M"), None
    first_line = str(value or "").split("\n")[0]
    m = re.search(r"(\d{1,2})\s*[:：]\s*(\d{2})(?:\s*[-–—~]\s*(\d{1,2})\s*[:：]\s*(\d{2}))?", first_line)
    if not m:
        return None, None
    start_min = int(m.group(1)) * 60 + int(m.group(2))
    start_time = f"{int(m.group(1)):02d}:{m.group(2)}"
    duration = None
    if m.group(3) is not None:
        end_min = int(m.group(3)) * 60 + int(m.group(4))
        duration = (end_min - start_min) % (24 * 60) or None
    return start_time, duration


@router.post("/import-excel")
async def import_plan_excel(node_id: int, file: UploadFile = File(...)):
    """从 Excel 导入学习计划。

    页签名按阶段编号（如 '1.2. 暑假第二周'）匹配该主题下的阶段并关联 phase_id；
    页签内每个非空格子导入为一条任务，日期取自列表头（如 '8月8日'），
    计划时间取自行首时间段（如 '9:00-11:00'）。
    幂等策略：先清除所涉及阶段下 source='excel' 的旧记录，再写入新任务。
    """
    db = await get_db()
    try:
        theme_phases = await query(db, "SELECT * FROM phases_v2 WHERE node_id = ?", (node_id,))
        if not theme_phases:
            return error("该主题下没有阶段，无法导入", 400)
        focus_items = await query(
            db,
            "SELECT id, title FROM planning_nodes WHERE parent_id = ? AND node_type = 'FOCUS_ITEM'",
            (node_id,))

        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)

        imported = 0
        sheet_results = []
        skipped = []
        cleared_phases = set()

        for ws in wb.worksheets:
            phase_num = _sheet_phase_number(ws.title)
            phase = next((p for p in theme_phases if str(p["phase_number"]) == phase_num), None) if phase_num else None
            if not phase:
                skipped.append(ws.title)
                continue
            try:
                phase_start = date_cls.fromisoformat(phase["start_date"])
                phase_end = date_cls.fromisoformat(phase["end_date"])
            except (TypeError, ValueError):
                skipped.append(f"{ws.title}（阶段日期无效）")
                continue

            # 定位表头行（A 列为“时间段”的那一行）
            header_row = None
            for r in range(1, min(ws.max_row, 5) + 1):
                if "时间段" in str(ws.cell(row=r, column=1).value or ""):
                    header_row = r
                    break
            if not header_row:
                skipped.append(f"{ws.title}（未找到表头）")
                continue

            # 列表头 -> 日期
            col_dates = {}
            for c in range(2, ws.max_column + 1):
                d = _parse_header_date(ws.cell(row=header_row, column=c).value, phase_start, phase_end)
                if d:
                    col_dates[c] = d
            if not col_dates:
                skipped.append(f"{ws.title}（未解析到日期列）")
                continue

            # 行首时间段 + 格子内容 -> 任务
            tasks = []
            for r in range(header_row + 1, ws.max_row + 1):
                start_time, duration = _parse_time_label(ws.cell(row=r, column=1).value)
                for c, exec_date in col_dates.items():
                    text = ws.cell(row=r, column=c).value
                    if text is None or not str(text).strip():
                        continue
                    title = str(text).strip()
                    # 任务文本以“科目：”开头时关联到对应重点项
                    task_node_id = phase["node_id"]
                    for f in focus_items:
                        if title.startswith(f["title"]):
                            task_node_id = f["id"]
                            break
                    tasks.append((task_node_id, exec_date, title, start_time, duration))

            # 幂等：清掉该阶段旧的 excel 导入记录
            if phase["id"] not in cleared_phases:
                await execute(db,
                              "DELETE FROM daily_executions WHERE phase_id = ? AND source = 'excel'",
                              (phase["id"],))
                cleared_phases.add(phase["id"])

            uid = phase["user_id"] or 1
            for task_node_id, exec_date, title, start_time, duration in tasks:
                await execute(
                    db,
                    """INSERT INTO daily_executions
                       (node_id, phase_id, user_id, execution_date, title,
                        planned_start_time, planned_duration, source,
                        is_done, completion_percent, result_score)
                       VALUES (?, ?, ?, ?, ?, ?, ?, 'excel', 0, 0, 0)""",
                    (task_node_id, phase["id"], uid, exec_date, title, start_time, duration),
                )
            imported += len(tasks)
            sheet_results.append({
                "sheet": ws.title,
                "phase": f"{phase['phase_number']}. {phase['title']}",
                "count": len(tasks),
            })

        await db.commit()
        if not sheet_results:
            return error("未匹配到任何阶段页签（页签名需以阶段编号开头，如「1.2. 暑假第二周」）", 400)
        msg = f"导入 {imported} 项任务"
        if skipped:
            msg += f"，跳过 {len(skipped)} 个页签"
        return success({"imported": imported, "sheets": sheet_results, "skipped": skipped}, msg)
    except Exception as e:
        return error(str(e), 500)
    finally:
        await db.close()


# ── Excel 计划导出 ──

@router.get("/export-excel/{phase_id}")
async def export_plan_excel(phase_id: int):
    """按导入模版的样式导出一个阶段的计划为 Excel 周计划表。

    页签名 = 阶段编号 + 标题（可被导入接口重新识别）；行 = 时间段、列 = 阶段内每一天，
    每个格子对应该日期该时间段的任务（多项任务以换行合并）。
    """
    db = await get_db()
    try:
        phase = await query_one(db, "SELECT * FROM phases_v2 WHERE id = ?", (phase_id,))
        if not phase:
            raise HTTPException(404, "阶段不存在")
        rows = await query(
            db,
            "SELECT * FROM daily_executions WHERE phase_id = ?"
            " ORDER BY execution_date, planned_start_time, id",
            (phase_id,),
        )
    finally:
        await db.close()

    try:
        start = date_cls.fromisoformat(phase["start_date"])
        end = date_cls.fromisoformat(phase["end_date"])
    except (TypeError, ValueError):
        raise HTTPException(400, "阶段日期格式无效")
    days = [start + timedelta(days=i) for i in range((end - start).days + 1)]

    def fmt_hm(total_min):
        return f"{total_min // 60}:{total_min % 60:02d}"

    # 按时间段分组：slot_key = (起始分钟, 时间段标签)；无时间的任务归入最后一行
    slots = {}   # key -> label
    cells = {}   # (slot_key, date) -> [titles]
    for r in rows:
        t = r["planned_start_time"]
        if t and re.match(r"^\d{1,2}:\d{2}", t):
            h, m = int(t[:2]), int(t[3:5])
            start_min = h * 60 + m
            label = fmt_hm(start_min)
            if r["planned_duration"]:
                label += f"-{fmt_hm((start_min + r['planned_duration']) % (24 * 60))}"
            key = (start_min, label)
        else:
            key = (24 * 60 + 1, "时间未定")
        slots.setdefault(key[0], key[1])
        cells.setdefault((key[0], r["execution_date"]), []).append(r["title"] or "")

    wb = openpyxl.Workbook()
    ws = wb.active
    sheet_name = re.sub(r"[\[\]:*?/\\]", "", f"{phase['phase_number']}. {phase['title']}")
    ws.title = sheet_name[:31]

    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="E8EEF7")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ncols = len(days) + 1
    # 第 1 行：标题（合并）
    title_text = f"{phase['title']}学习日程表（{start.month}月{start.day}日 - {end.month}月{end.day}日）"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=title_text)
    c.font = Font(bold=True, size=14)
    c.alignment = center
    # 第 2 行：表头（时间段 + 日期列）
    weekday_cn = "一二三四五六日"
    c = ws.cell(row=2, column=1, value="时间段")
    c.font = Font(bold=True)
    c.fill = header_fill
    c.alignment = center
    c.border = border
    for i, d in enumerate(days):
        c = ws.cell(row=2, column=2 + i, value=f"{d.month}月{d.day}日\n周{weekday_cn[d.weekday()]}")
        c.font = Font(bold=True)
        c.fill = header_fill
        c.alignment = center
        c.border = border
    # 数据行
    for ri, start_min in enumerate(sorted(slots), start=3):
        c = ws.cell(row=ri, column=1, value=slots[start_min])
        c.alignment = center
        c.border = border
        for i, d in enumerate(days):
            texts = cells.get((start_min, d.isoformat()))
            c = ws.cell(row=ri, column=2 + i, value="\n".join(texts) if texts else None)
            c.alignment = left_wrap
            c.border = border

    ws.column_dimensions[get_column_letter(1)].width = 14
    for i in range(len(days)):
        ws.column_dimensions[get_column_letter(2 + i)].width = 32

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = quote(f"{phase['phase_number']}.{phase['title']}计划.xlsx")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )


# ── 当日计划打印导出 ──

@router.get("/export-daily/{node_id}")
async def export_daily_plan(node_id: int, date: str = None):
    """导出某一天的计划为可打印 Excel 表格。

    列：计划时间 | 任务内容 | 实际开始时间 | 实际结束时间 | 时长 | 其它说明，
    其中后 4 列留空，供打印后手工填写。date 缺省为今天。
    """
    day = date or date_cls.today().isoformat()
    try:
        day_obj = date_cls.fromisoformat(day)
    except ValueError:
        raise HTTPException(400, "日期格式无效")

    db = await get_db()
    try:
        theme = await query_one(db, "SELECT * FROM planning_nodes WHERE id = ?", (node_id,))
        if not theme:
            raise HTTPException(404, "主题不存在")
        rows = await query(
            db,
            """WITH RECURSIVE subtree AS (
                   SELECT id FROM planning_nodes WHERE id = ?
                   UNION ALL
                   SELECT n.id FROM planning_nodes n
                   JOIN subtree s ON n.parent_id = s.id
                 )
                 SELECT * FROM daily_executions
                 WHERE node_id IN (SELECT id FROM subtree) AND execution_date = ?
                 ORDER BY CASE WHEN planned_start_time IS NULL OR planned_start_time = '' THEN 1 ELSE 0 END,
                          planned_start_time, id""",
            (node_id, day),
        )
    finally:
        await db.close()

    def fmt_hm(total_min):
        return f"{total_min // 60:02d}:{total_min % 60:02d}"

    def planned_label(r):
        t = r["planned_start_time"]
        if not (t and re.match(r"^\d{1,2}:\d{2}", t)):
            return "时间未定"
        label = t[:5]
        if r["planned_duration"]:
            start_min = int(t[:2]) * 60 + int(t[3:5])
            label += f"-{fmt_hm((start_min + r['planned_duration']) % (24 * 60))}"
        return label

    wb = openpyxl.Workbook()
    ws = wb.active
    weekday_cn = "一二三四五六日"
    ws.title = f"{day_obj.month}月{day_obj.day}日计划"

    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
    header_fill = PatternFill("solid", fgColor="E8EEF7")

    headers = ["计划时间", "任务内容", "实际开始时间", "实际结束时间", "时长", "其它说明"]
    ncols = len(headers)
    # 第 1 行：标题（合并）
    title_text = (f"{theme['title']} · {day_obj.month}月{day_obj.day}日"
                  f"（周{weekday_cn[day_obj.weekday()]}）学习计划")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=title_text)
    c.font = Font(bold=True, size=14)
    c.alignment = center
    ws.row_dimensions[1].height = 28
    # 第 2 行：表头
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = Font(bold=True)
        c.fill = header_fill
        c.alignment = center
        c.border = border
    ws.row_dimensions[2].height = 22
    # 数据行：仅填计划时间/任务内容，后 4 列留空供手工填写
    for ri, r in enumerate(rows, start=3):
        c = ws.cell(row=ri, column=1, value=planned_label(r))
        c.alignment = center
        c.border = border
        c = ws.cell(row=ri, column=2, value=r["title"] or "")
        c.alignment = left_wrap
        c.border = border
        for ci in range(3, ncols + 1):
            c = ws.cell(row=ri, column=ci, value=None)
            c.alignment = center
            c.border = border
        ws.row_dimensions[ri].height = 26  # 行高留出手写空间

    for i, w in enumerate([14, 42, 14, 14, 10, 26], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    # 打印设置：A4 纵向、按宽度缩放一页
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = quote(f"{day}当日计划.xlsx")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )


# ── File Upload ──

@router.post("/daily-executions/{exec_id}/upload")
async def upload_execution_file(exec_id: int, file: UploadFile = File(...)):
    """上传执行记录附件到 execution-docs/ 目录"""
    try:
        upload_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "uploads", "execution-docs")
        os.makedirs(upload_dir, exist_ok=True)

        # Generate unique filename: exec_{id}_{timestamp}_{original}
        ts = datetime.now().strftime("%Y%m%d%H%M%S")
        safe_name = re.sub(r'[^\w.\-]', '_', file.filename or "file")
        filename = f"exec_{exec_id}_{ts}_{safe_name}"
        filepath = os.path.join(upload_dir, filename)

        content = await file.read()
        with open(filepath, "wb") as f:
            f.write(content)

        # Store URL in database
        db = await get_db()
        try:
            row = await query_one(db, "SELECT images FROM daily_executions WHERE id = ?", (exec_id,))
            existing = row["images"] or "" if row else ""
            new_images = (existing + "," + filename) if existing else filename
            await execute(db, "UPDATE daily_executions SET images = ? WHERE id = ?", (new_images, exec_id))
        finally:
            await db.close()

        return success({"filename": filename, "url": f"/uploads/execution-docs/{filename}"}, "上传成功")
    except Exception as e:
        return error(str(e), 500)
