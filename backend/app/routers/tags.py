"""Tag routes (统一标签库：标签类型 + 标签，全局通用)."""
import io
import os
from urllib.parse import quote

import openpyxl
from fastapi import APIRouter, Depends, UploadFile, File
from fastapi.responses import FileResponse, StreamingResponse

from app.database import get_db, query, query_one, execute
from app.models import success, error
from app.models.tag import TagCreate, TagUpdate, TagTypeCreate

router = APIRouter(prefix="/api/tags", tags=["标签"])


def _parse_color(v):
    s = str(v or "").strip()
    return s or None


# ── 标签类型 ──

@router.get("/types")
async def list_tag_types():
    """列出所有标签类型（kind='type'）。"""
    db = await get_db()
    try:
        rows = await query(
            db,
            "SELECT * FROM tags WHERE kind = 'type' ORDER BY sort_order, id",
        )
        return success(rows)
    finally:
        await db.close()


@router.post("/types")
async def create_tag_type(body: TagTypeCreate):
    """新增标签类型。"""
    db = await get_db()
    try:
        name = body.name.strip()
        if not name:
            return error("类型名不能为空", 400)
        existing = await query_one(
            db, "SELECT id FROM tags WHERE kind = 'type' AND name = ?", (name,)
        )
        if existing:
            return error("该标签类型已存在", 400)
        last_id, _ = await execute(
            db,
            "INSERT INTO tags (name, kind, sort_order) VALUES (?, 'type', (SELECT COALESCE(MAX(sort_order),0)+1 FROM tags WHERE kind='type'))",
            (name,),
        )
        row = await query_one(db, "SELECT * FROM tags WHERE id = ?", (last_id,))
        return success(row, "标签类型创建成功")
    finally:
        await db.close()


@router.delete("/types/{type_id}")
async def delete_tag_type(type_id: int):
    """删除标签类型，并级联删除该类型下的所有标签。"""
    db = await get_db()
    try:
        row = await query_one(db, "SELECT * FROM tags WHERE id = ? AND kind = 'type'", (type_id,))
        if not row:
            return error("标签类型不存在", 404)
        # 级联删除该类型下所有标签（含其子标签）
        await execute(db, "DELETE FROM tags WHERE type_name = ? AND kind = 'tag'", (row["name"],))
        await execute(db, "DELETE FROM tags WHERE id = ?", (type_id,))
        return success({"deleted_id": type_id}, "标签类型删除成功")
    finally:
        await db.close()


# ── 标签 ──

@router.get("")
async def list_tags(type_name: str = None, page: int = 1, size: int = 20):
    """标签列表（树形：一级标签含 children），可按 type_name 过滤，按一级标签分页。"""
    db = await get_db()
    try:
        page = max(page, 1)
        size = max(size, 1)
        where = ["t.kind = 'tag'", "t.parent_id IS NULL"]
        params = []
        if type_name:
            where.append("t.type_name = ?")
            params.append(type_name)

        # 一级标签总数
        total_row = await query_one(
            db,
            f"SELECT COUNT(*) AS c FROM tags t WHERE {' AND '.join(where)}",
            tuple(params),
        )
        total = total_row["c"] if total_row else 0

        # 分页查询一级标签
        offset = (page - 1) * size
        top_rows = await query(
            db,
            f"""SELECT t.*, n.title AS focus_title
                FROM tags t
                LEFT JOIN planning_nodes n ON t.focus_id = n.id
                WHERE {' AND '.join(where)}
                ORDER BY t.type_name, t.sort_order, t.id
                LIMIT ? OFFSET ?""",
            tuple(params) + (size, offset),
        )

        # 查询当前页一级标签的所有二级标签
        children = []
        top_ids = [r["id"] for r in top_rows]
        if top_ids:
            ph = ",".join("?" for _ in top_ids)
            children = await query(
                db,
                f"""SELECT t.*, n.title AS focus_title
                    FROM tags t
                    LEFT JOIN planning_nodes n ON t.focus_id = n.id
                    WHERE t.kind = 'tag' AND t.parent_id IN ({ph})
                    ORDER BY t.sort_order, t.id""",
                tuple(top_ids),
            )

        # 组装树形
        items = []
        for top in top_rows:
            item = dict(top)
            item["children"] = [c for c in children if c["parent_id"] == top["id"]]
            items.append(item)

        total_pages = (total + size - 1) // size if total else 0
        return success({"tags": items, "total": total, "page": page, "size": size, "total_pages": total_pages})
    finally:
        await db.close()


@router.post("")
async def create_tag(body: TagCreate):
    db = await get_db()
    try:
        name = body.name.strip()
        if not name:
            return error("标签名不能为空", 400)
        # 校验父标签存在性，避免自引用外键约束失败（前端列表可能未及时刷新）
        if body.parent_id is not None:
            parent = await query_one(db, "SELECT id, kind FROM tags WHERE id = ?", (body.parent_id,))
            if not parent or parent["kind"] != "tag":
                return error("父标签不存在或已删除，请刷新列表后重试", 400)
        last_id, _ = await execute(
            db,
            "INSERT INTO tags (name, kind, type_name, parent_id, color, description, focus_id) VALUES (?, 'tag', ?, ?, ?, ?, ?)",
            (name, body.type_name, body.parent_id, body.color, body.description, body.focus_id),
        )
        row = await query_one(db, "SELECT * FROM tags WHERE id = ?", (last_id,))
        return success(row, "标签创建成功")
    finally:
        await db.close()


@router.put("/{tag_id}")
async def update_tag(tag_id: int, body: TagUpdate):
    db = await get_db()
    try:
        data = body.model_dump(exclude_unset=True)
        if "name" in data and data["name"] is not None:
            data["name"] = data["name"].strip()
            if not data["name"]:
                return error("标签名不能为空", 400)
        # 校验父标签存在性（含防止把标签设置成自己的父标签形成环）
        if "parent_id" in data and data["parent_id"] is not None:
            if data["parent_id"] == tag_id:
                return error("父标签不能是自身", 400)
            parent = await query_one(db, "SELECT id, kind FROM tags WHERE id = ?", (data["parent_id"],))
            if not parent or parent["kind"] != "tag":
                return error("父标签不存在或已删除，请刷新列表后重试", 400)
        fields = []
        values = []
        for key in ["name", "type_name", "parent_id", "color", "description", "focus_id"]:
            if key in data:
                fields.append(f"{key} = ?")
                values.append(data[key])
        if fields:
            fields.append("updated_at = CURRENT_TIMESTAMP")
            values.append(tag_id)
            await execute(db, f"UPDATE tags SET {', '.join(fields)} WHERE id = ?", tuple(values))
        row = await query_one(db, "SELECT * FROM tags WHERE id = ?", (tag_id,))
        if not row:
            return error("标签不存在", 404)
        return success(row, "标签更新成功")
    finally:
        await db.close()


@router.delete("/{tag_id}")
async def delete_tag(tag_id: int):
    db = await get_db()
    try:
        # 删除该标签时，其下二级标签一并删除（外键级联）
        await execute(db, "DELETE FROM tags WHERE parent_id = ?", (tag_id,))
        _, changes = await execute(db, "DELETE FROM tags WHERE id = ? AND kind = 'tag'", (tag_id,))
        if changes == 0:
            return error("标签不存在", 404)
        return success({"deleted_id": tag_id}, "标签删除成功")
    finally:
        await db.close()


# ── Excel 标签导入/模板下载 ──

@router.get("/export")
async def export_tags(type_name: str = None):
    """导出标签为 Excel（列与导入模版一致：标签类型/一级标签/二级标签/说明/颜色）。

    可按 type_name 过滤，与标签管理页当前查询结果一致。
    """
    db = await get_db()
    try:
        where = ["t.kind = 'tag'"]
        params = []
        if type_name:
            where.append("t.type_name = ?")
            params.append(type_name)
        rows = await query(
            db,
            f"""SELECT t.*, p.name AS parent_name, n.title AS focus_title
                FROM tags t
                LEFT JOIN tags p ON t.parent_id = p.id
                LEFT JOIN planning_nodes n ON t.focus_id = n.id
                WHERE {' AND '.join(where)}
                ORDER BY t.type_name, t.sort_order, t.id""",
            tuple(params),
        )
    finally:
        await db.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "标签"
    ws.append(["标签类型", "一级标签", "二级标签", "说明", "颜色", "关联重点"])
    for row in rows:
        ws.append([
            row["type_name"] or "",
            row["parent_name"] or row["name"],
            row["name"] if row["parent_id"] else "",
            row["description"] or "",
            row["color"] or "",
            row["focus_title"] or "",
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = quote(f"标签导出-{type_name}.xlsx" if type_name else "标签导出.xlsx")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )


@router.get("/template/download")
async def download_tag_template():
    """下载标签导入模版（静态文件）。"""
    root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__))))
    path = os.path.join(root, "docs", "design", "标签导入模版.xlsx")
    if not os.path.exists(path):
        return error("模版文件不存在", 404)
    return FileResponse(
        path,
        filename="标签导入模版.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


async def _ensure_tag_type(db, type_name) -> None:
    """确保标签类型存在，不存在则新增。"""
    type_name = type_name.strip()
    if not type_name:
        return
    existing = await query_one(
        db, "SELECT id FROM tags WHERE kind = 'type' AND name = ?", (type_name,)
    )
    if existing:
        return
    await execute(
        db,
        "INSERT INTO tags (name, kind, sort_order) VALUES (?, 'type', (SELECT COALESCE(MAX(sort_order),0)+1 FROM tags WHERE kind='type'))",
        (type_name,),
    )


async def _find_or_create_tag(db, name, type_name, parent_id, color, description, focus_id=None) -> int:
    """按 (name, type_name, parent_id) 查找标签，不存在则创建，返回标签 id。"""
    if parent_id is None:
        row = await query_one(
            db,
            "SELECT id FROM tags WHERE kind='tag' AND name = ? AND type_name = ? AND parent_id IS NULL",
            (name, type_name),
        )
    else:
        row = await query_one(
            db,
            "SELECT id FROM tags WHERE kind='tag' AND name = ? AND type_name = ? AND parent_id = ?",
            (name, type_name, parent_id),
        )
    if row:
        # 已存在时，若传入了关联重点则更新
        if focus_id is not None:
            await execute(db, "UPDATE tags SET focus_id = ? WHERE id = ?", (focus_id, row["id"]))
        return row["id"]
    last_id, _ = await execute(
        db,
        "INSERT INTO tags (name, kind, type_name, parent_id, color, description, focus_id) VALUES (?, 'tag', ?, ?, ?, ?, ?)",
        (name, type_name, parent_id, color, description, focus_id),
    )
    return last_id


@router.post("/import-excel")
async def import_tag_excel(file: UploadFile = File(...)):
    """从 Excel 导入标签。

    列：A=标签类型 B=一级标签 C=二级标签 D=说明 E=颜色 F=关联重点
    标签类型不存在时自动新增；同名标签不重复创建；关联重点按名称匹配重点项。
    """
    db = await get_db()
    try:
        contents = await file.read()
        try:
            wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        except Exception:
            return error("无法解析 Excel 文件", 400)

        ws = wb.worksheets[0] if wb.worksheets else None
        if not ws:
            return error("Excel 文件为空", 400)

        # 预加载重点项名称 -> id 映射
        focus_items = await query(db, "SELECT id, title FROM planning_nodes WHERE node_type = 'FOCUS_ITEM'")
        focus_map = {f["title"]: f["id"] for f in focus_items}

        imported = 0
        created_types = []
        # 定位表头行
        header_row = None
        for r in range(1, min(ws.max_row, 5) + 1):
            a = str(ws.cell(row=r, column=1).value or "")
            b = str(ws.cell(row=r, column=2).value or "")
            if "类型" in a or "标签" in b or "一级" in b:
                header_row = r
                break
        if not header_row:
            header_row = 1

        for r in range(header_row + 1, ws.max_row + 1):
            type_name = str(ws.cell(row=r, column=1).value or "").strip()
            first = str(ws.cell(row=r, column=2).value or "").strip()
            second = str(ws.cell(row=r, column=3).value or "").strip()
            description = str(ws.cell(row=r, column=4).value or "").strip() or None
            color = _parse_color(ws.cell(row=r, column=5).value)
            focus_name = str(ws.cell(row=r, column=6).value or "").strip()
            focus_id = focus_map.get(focus_name) if focus_name else None

            if not first:
                continue  # 空行跳过
            if not type_name:
                continue  # 缺标签类型跳过

            # 确保标签类型存在
            existed = await query_one(db, "SELECT id FROM tags WHERE kind='type' AND name = ?", (type_name,))
            if not existed:
                await _ensure_tag_type(db, type_name)
                created_types.append(type_name)

            # 一级标签
            first_id = await _find_or_create_tag(db, first, type_name, None, color, description)
            # 二级标签（可选）
            if second:
                await _find_or_create_tag(db, second, type_name, first_id, color, description, focus_id)
            elif focus_id is not None:
                await execute(db, "UPDATE tags SET focus_id = ? WHERE id = ?", (focus_id, first_id))
            imported += 1

        await db.commit()
        msg = f"导入 {imported} 个标签"
        if created_types:
            msg += f"，新增标签类型 {len(created_types)} 个"
        return success({"imported": imported, "created_types": created_types}, msg)
    except Exception as e:
        return error(str(e), 500)
    finally:
        await db.close()
